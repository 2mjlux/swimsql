# This Source Code Form is subject to the terms of the Mozilla Public License, v. 2.0.
# If a copy of the MPL was not distributed with this file, You can obtain one at
# http://mozilla.org/MPL/2.0/.
# Copyright (c) 2026 Michael JJ Martin (https://github.com/2mjlux)


import db
from pathlib import Path
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from openpyxl import Workbook
from openpyxl.styles import Font


HEADERS_PERFORMANCES_METRES = [
    "Swimmer", "Pool", "Meet", "Meet Start",
    "Discipline", "Time", "Date", "Session", "Notes"
]

HEADERS_PERFORMANCES_YARDS = [
    "Swimmer", "Meet", "Meet Start",
    "Discipline", "Time", "Date", "Session", "Notes"
]

HEADERS_PERSONAL_BESTS_METRES = [
    "Swimmer", "Discipline", "Pool", "Best Time", "Date"
]

HEADERS_PERSONAL_BESTS_YARDS = [
    "Swimmer", "Discipline", "Best Time", "Date"
]


def _add_sheet_ods(doc, sheet_name, headers, rows):
    """
    Add a sheet to an ODS document with bold headers and data rows.
    """
    sheet = Table(name=sheet_name)
    doc.spreadsheet.addElement(sheet)

    # Header row
    header_row = TableRow()
    for header in headers:
        cell = TableCell()
        cell.addElement(P(text=header))
        header_row.addElement(cell)
    sheet.addElement(header_row)

    # Data rows
    for row in rows:
        data_row = TableRow()
        for value in row:
            cell = TableCell()
            cell.addElement(P(text=str(value) if value is not None else ""))
            data_row.addElement(cell)
        sheet.addElement(data_row)


def _add_sheet_xlsx(wb, sheet_name, headers, rows):
    """
    Add a sheet to an XLSX workbook with bold headers and data rows.
    """
    ws = wb.create_sheet(title=sheet_name)

    # Bold header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for row in rows:
        ws.append(row)


def _get_export_data():
    """
    Fetch all data needed for export from db.py.
    Returns a dictionary with four keys.
    """
    return {
        "performances_metres": db.list_all_performances_metres(),
        "performances_yards":  db.list_all_performances_yards(),
        "personal_bests_metres": db.get_all_personal_bests_metres(),
        "personal_bests_yards":  db.get_all_personal_bests_yards(),
    }


def export_ods(filepath):
    """
    Export all data to ODS format.
    Four sheets: Performances Metres, Performances Yards,
    Personal Bests Metres, Personal Bests Yards.
    """
    doc = OpenDocumentSpreadsheet()  # create the document
    data = _get_export_data()  # fetch the data

    # call _add_sheet_ods() four times - once per sheet
    # Sheet 1 - Performance Metres
    rows = [
        [
            row["swimmer"],
            row["pool"],
            row["meet"],
            row["date_start"],
            row["discipline"],
            db.cs_to_time(row["time_cs"]),
            row["date"],
            row["session"] or "",
            row["notes"] or "",
        ]
        for row in data["performances_metres"]
    ]
    _add_sheet_ods(doc, "Performances Metres", HEADERS_PERFORMANCES_METRES, rows)

    # Sheet 2 - Performance Yards
    rows = [
        [
            row["swimmer"],
            row["meet"],
            row["date_start"],
            row["discipline"],
            db.cs_to_time(row["time_cs"]),
            row["date"],
            row["session"] or "",
            row["notes"] or "",
        ]
        for row in data["performances_yards"]
    ]
    _add_sheet_ods(doc, "Performances Yards", HEADERS_PERFORMANCES_YARDS, rows)

    # Sheet 3 - Personal Bests Metres
    rows = [
        [
            row["swimmer"],
            row["discipline"],
            row["pool"],
            db.cs_to_time(row["best_cs"]),
            row["date"],
        ]
        for row in data["personal_bests_metres"]
    ]
    _add_sheet_ods(doc, "Personal Bests Metres", HEADERS_PERSONAL_BESTS_METRES, rows)

    # Sheet 4 - Personal Bests Yards
    rows = [
        [
            row["swimmer"],
            row["discipline"],
            db.cs_to_time(row["best_cs"]),
            row["date"],
        ]
        for row in data["personal_bests_yards"]
    ]
    _add_sheet_ods(doc, "Personal Bests Yards", HEADERS_PERSONAL_BESTS_YARDS, rows)

    # save the document
    doc.save(filepath)


def export_xlsx(filepath):
    """
    Export all data to XLSX format.
    Four sheets: Performances Metres, Performances Yards,
    Personal Bests Metres, Personal Bests Yards.
    """
    wb = Workbook()  # create the workbook
    wb.remove(wb.active)  # remove empty default sheet before adding ours
    data = _get_export_data()  # fetch the data

    # call _add_sheet_xlsx() four times - once per sheet
    # Sheet 1 - Performance Metres
    rows = [
        [
            row["swimmer"],
            row["pool"],
            row["meet"],
            row["date_start"],
            row["discipline"],
            db.cs_to_time(row["time_cs"]),
            row["date"],
            row["session"] or "",
            row["notes"] or "",
        ]
        for row in data["performances_metres"]
    ]
    _add_sheet_xlsx(wb, "Performances Metres", HEADERS_PERFORMANCES_METRES, rows)

    # Sheet 2 - Performance Yards
    rows = [
        [
            row["swimmer"],
            row["meet"],
            row["date_start"],
            row["discipline"],
            db.cs_to_time(row["time_cs"]),
            row["date"],
            row["session"] or "",
            row["notes"] or "",
        ]
        for row in data["performances_yards"]
    ]
    _add_sheet_xlsx(wb, "Performances Yards", HEADERS_PERFORMANCES_YARDS, rows)

    # Sheet 3 - Personal Bests Metres
    rows = [
        [
            row["swimmer"],
            row["discipline"],
            row["pool"],
            db.cs_to_time(row["best_cs"]),
            row["date"],
        ]
        for row in data["personal_bests_metres"]
    ]
    _add_sheet_xlsx(wb, "Personal Bests Metres", HEADERS_PERSONAL_BESTS_METRES, rows)

    # Sheet 4 - Personal Bests Yards
    rows = [
        [
            row["swimmer"],
            row["discipline"],
            db.cs_to_time(row["best_cs"]),
            row["date"],
        ]
        for row in data["personal_bests_yards"]
    ]
    _add_sheet_xlsx(wb, "Personal Bests Yards", HEADERS_PERSONAL_BESTS_YARDS, rows)

    # save the document
    wb.save(filepath)
